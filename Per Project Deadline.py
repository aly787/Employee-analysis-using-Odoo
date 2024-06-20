from openpyxl import load_workbook
from datetime import datetime
from dateutil.parser import parser

CurrentDate = datetime.now()
wb = load_workbook("Task (project.task) (7).xlsx")
sheet = wb["Sheet1"]
all_companies = []
current_company = {"Project": "",
                   "Task Owner": "",
                   "Assignee": "",
                   "Planned Hours": "",
                   "Remaining Hours": "",
                   "Status": "",
                   "Stage": "",
                   "Deadline": "",
                   "prodution": "",
                   "support":""  }

#Current: Trying to convert the deadline date in string format to a proper date and compare it to current date to check if its late
# late_companies = []
# proj = ""
# owner = ""
# assign = ""
# phours = ""
# rhours = ""
# stat = ""
# stag = ""
# dead = ""

for row in sheet.iter_rows(min_row=2):
    if row[0].value is not None:
        if (row[3].value is not None):
            if ("maha" in row[3].value):
                if "SUPPORT" in row[8].value:
                    d1 = {"Project": row[0].value,
                            "Task Owner": row[2].value,
                            "Assignee": row[3].value,
                            "Planned Hours": row[4].value,
                            "Remaining Hours": row[5].value,
                            "Status": row[7].value,
                            "Stage": row[8].value,
                            "Deadline": row[9].value,
                            "production": row[12].value,
                            "support": row[13].value  }
                    print(d1)
                    print("\n")


#         if (row[9].value) is not None:
#             datetime_obj = parser().parse(str(row[9].value), yearfirst = True)
#             if datetime_obj > CurrentDate:
#                 d1.update({"Late": "True"})
#         else :
#             d1.update({"Late": "False"})
#         all_companies.append(d1)
# for n in all_companies:
#     for key, value in n.items():
#         if (key == "Late") and (value == "True"):
#             p = n["Project"]
#             t = n["Task Owner"]
#             a = n["Assignee"]
#             if (p == None):
#                 p = ""
#             if (t == None):
#                 t = ""
#             if (a == None):
#                 a = ""
#             print(p + "," + t + "," + a)


#print(all_companies)
    #if ("Issue" in row[6].value) and (("FIG" in row[1].value) or ("FIG" in row[2].value))  :
    #String here is the comparison value to check what is in the row at column N

    # if "issue" in row[6].value :
    #     print(current_company)

#Other Tests:
    # original = row[6].value
    # original1 = row[1].value
    # original2 = row[2].value
    # original3 = row[3].value

    # if row[0].value is not None:
    #     current_owner = row[1].value

    #     if original1 is None:
    #         current_owner = "none"
    #     if original3 is None:
    #         current_subtasks = 0
    #     if original2 is None:
    #         current_assignee = "none"
    #     current_assignee = row[2].value
    #     current_company = row[0].value
    #     current_subtasks = row[3].value

    # if original is None:
    #     continue
    
    # if ("industry" in row[6].value):
    #     counter += 1
    

    # #String here is the comparison value to check what is in the row at column N
    #     print(current_company + "->" + current_owner + "->" + current_assignee + "->    " + str(current_subtasks))

    