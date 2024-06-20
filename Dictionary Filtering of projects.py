from ctypes import sizeof
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from dateutil.parser import parser

CurrentDate = datetime.now()
wb = load_workbook("Task (project.task) (13).xlsx")
sheet = wb["Sheet1"]

#List of all the projects as dictionaries
all_companies = []
d1 = {"Project": "",
        "Type": "",
        "Task Owner": "",
        "Assignee": [],
        "Tags": [],
        "Planned Hours": "",
        "Hours Spent": "",
        "Stage": "",
        "Creation": "",
        "Assigning Date": "",
        "Date in production": "",
        "Date in support": "",
                }
#Dictionary creation with all data for each project including the multiple assignees
for row in sheet.rows:
    if row[0].value is not None:
        d1 = {"Project": row[0].value,
                "Type": row[1].value,
                "Task Owner": row[2].value,
                "Planned Hours": row[5].value,
                "Hours Spent": row[6].value,
                "Stage": row[7].value,
                "Creation": row[8].value,
                "Assigning Date": row[9].value,
                "Date in production": row[10].value,
                "Date in support": row[11].value,
                }
        all_companies.append(d1)
    # d1["Assignee"].append(row[3].value)
    # d1["Tags"].append(row[4].value)
    # if row[0] is None:
    #     for n in range(12):
    #         if row[n] is not None:
    #             if (n == 3) and (row[3].value is not None):
    #                 place = d1["Assignee"][0]
    #                 newPlace = str(place) + ", " + str(row[3].value)
    #                 d1["Assignee"] = newPlace
    #             if (n == 4) and (row[3].value is not None):
    #                 place = d1["Tags"][0]
    #                 newPlace = str(place) + ", " + str(row[4].value)
    #                 d1["Tags"] = newPlace
    # if row[0].value is None:
    #     current_dict = all_companies[len(all_companies) - 1]
    #     current_dict["Assignee"].append(row[3].value)
    #     del all_companies[-1]
    #     all_companies.append(current_dict)
print(all_companies)

workbook = Workbook()
ws = workbook.active
ws.title = "Projects"
counter = 2
for company in all_companies:
    ws["B{}".format(counter)] = list(company.items())[0][1] #Project Name
    ws["C{}".format(counter)] = list(company.items())[1][1] #Project Categories
    ws["D{}".format(counter)] = list(company.items())[2][1] #Task Owner
    ws["E{}".format(counter)] = list(company.items())[3][1] #Planned Hours
    ws["F{}".format(counter)] = list(company.items())[4][1] #Hours Spent
    ws["G{}".format(counter)] = list(company.items())[5][1] #Stage
    ws["H{}".format(counter)] = list(company.items())[6][1] #Task Created
    ws["I{}".format(counter)] = list(company.items())[7][1] #Assigning date
    ws["J{}".format(counter)] = list(company.items())[8][1] #Date in production
    ws["K{}".format(counter)] = list(company.items())[9][1] #Date in support
    counter += 1
workbook.save("/Users/aliottoman/Documents/Odoo/Test/Task (project.task) (6).xlsx")